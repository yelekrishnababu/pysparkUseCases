from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
wb=Workbook()
filePath="/home/ky2910/PycharmProjects/pysparkUseCases/inputs/test.xlsx"
wb=load_workbook(filePath)
print(wb.sheetnames)