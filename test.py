from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from pyspark.sql import SparkSession, DataFrameWriter

jdbcHostname = "sqlserver045.database.windows.net"
jdbcPort = "1433"
jdbcDatabase = "ky2910"
properties = {
 "user" : "ky2910@sqlserver045",
 "password" : "Krishna@2000" }
url = "jdbc:sqlserver://{0}:{1};database={2}".format(jdbcHostname,jdbcPort,jdbcDatabase)
spark = SparkSession.builder.master("local[1]").appName('SparkByExamples.com').getOrCreate()
spark.conf.set("spark.sql.execution.arrow.pyspark.enabled", "true")

wb=Workbook()
filePath="/home/ky2910/PycharmProjects/pysparkUseCases/inputs/ky2910.xlsx"
wb=load_workbook(filePath)
for i in wb.sheetnames:
    print(i)
    df = spark.read.format("com.crealytics.spark.excel") \
            .option("header", "true") \
            .option("treatEmptyValuesAsNulls", "false")\
            .option("dataAddress", f"'{i}'!A1:P20") \
        .option("treatEmptyValuesAsNulls", "false")\
            .load(filePath)
    # myfinaldf = DataFrameWriter(df)
    # myfinaldf.jdbc(url="jdbc:sqlserver://sqlserver045.database.windows.net:1433;database=ky2910;user=ky2910@sqlserver045;password={your_password_here};encrypt=true;trustServerCertificate=false;hostNameInCertificate=*.database.windows.net;loginTimeout=30;", table= f'{i}_ky2910', mode ="overwrite", properties = properties)
    df.write.format("jdbc").mode("overwrite").option("url","jdbc:postgresql://w3.training5.modak.com:5432/training")\
        .option("driver", "org.postgresql.Driver").option("dbtable", f'{i}KY2910')\
        .option("user", "mt4038").option("password","mt4038@m02y22").save()